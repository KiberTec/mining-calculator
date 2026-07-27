#!/usr/bin/env python3
"""Generate Excel with mining journal formulas — fixed cell references."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT = "/Users/kiber/Desktop/Mining Calculator/mining_calculator_formulas.xlsx"

INPUT_FILL = PatternFill("solid", fgColor="FFF8E7")
CALC_FILL = PatternFill("solid", fgColor="E8F4FD")
SECTION_FILL = PatternFill("solid", fgColor="F7931A")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Fixed row map (column B)
R = {
    "month": 4, "loc": 5, "kwh": 6, "rub": 7, "hash": 8, "diff": 9,
    "btc": 10, "given": 11, "sold": 12, "price": 13, "rate": 14,
    "cur_btc": 15, "cur_rub": 16,
    "days": 19, "hours": 20, "pow32": 21, "reward": 22,
    "elec_usd": 23, "comm_usd": 24, "cost_usd": 25, "cost_rub": 26,
    "net_btc": 27, "held": 28, "theor": 29,
    "rev_rub": 31, "hodl_rub": 32, "mtm": 33, "real_cash": 34, "fin": 35,
    "cost_btc_usd": 36, "cost_btc_rub": 37, "tariff": 38,
    "jth": 40, "real_pct": 41, "margin": 42, "breakeven": 43,
    "hp_our": 44, "hp_net": 45, "hp_diff": 46,
}

B = lambda key: f"B{R[key]}"


def section(ws, row, text, cols=3):
    ws.cell(row, 1, text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = BORDER


def put(ws, row, label, value, is_input=False, comment=None, fmt=None):
    ws.cell(row, 1, label).border = BORDER
    ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    c = ws.cell(row, 2, value)
    c.fill = INPUT_FILL if is_input else CALC_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        c.number_format = fmt
    if comment:
        c.comment = Comment(comment, "Mining Calculator", width=300, height=100)
        ws.cell(row, 3, comment).alignment = Alignment(wrap_text=True, vertical="top")


def build_calculator(ws):
    ws.title = "Расчёт месяца"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 52

    ws["A1"] = "Калькулятор записи журнала — меняй жёлтые ячейки (колонка B)"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:C1")
    ws.cell(2, 3, "Пояснение").font = Font(bold=True)

    section(ws, 3, "📥 ВВОД (жёлтые ячейки)")
    put(ws, 4, "Месяц (YYYY-MM)", "2026-04", True, "Для расчёта дней в месяце")
    put(ws, 5, "Локация", "Дема", True, "Бирск или Дема (инфо, на формулы не влияет)")
    put(ws, 6, "Потребление за месяц, кВт·ч", 89680, True, "Факт с счётчика за месяц")
    put(ws, 7, "Заплачено за эл-во, ₽", 325650, True, "Сколько заплатил за электричество")
    put(ws, 8, "Средний хешрейт, TH/s", 4500, True, "Среднее с пула за месяц (не пик!)")
    put(ws, 9, "Сложность сети, T", 135.59, True, "Сложность на ~28-е число месяца", fmt="0.00")
    put(ws, 10, "Намайнено, BTC (gross)", 0.06174, True, "Факт с пула", fmt="0.00000")
    put(ws, 11, "Отдано партнёру, BTC", 0, True, "Только Дема — BTC хостингу", fmt="0.00000")
    put(ws, 12, "Продано, BTC", 0.01, True, "Сколько продал", fmt="0.00000")
    put(ws, 13, "Цена продажи, $/BTC", 69000, True, "Цена продажи", fmt="#,##0")
    put(ws, 14, "Курс RUB при продаже", 79, True, "₽ за $1", fmt="0.00")
    put(ws, 15, "Текущая цена BTC, $", 63000, True, "С банера сайта — для HODL", fmt="#,##0")
    put(ws, 16, "Текущий курс USD/RUB", 78, True, "С банера сайта", fmt="0.00")

    section(ws, 18, "⚙️ БАЗОВЫЕ РАСЧЁТЫ")
    put(ws, 19, "Дней в месяце", f'=DAY(EOMONTH(DATEVALUE({B("month")}&"-01"),0))',
        comment="Авто из месяца")
    put(ws, 20, "Часов в месяце", f"={B('days')}*24", comment="Дни × 24")
    put(ws, 21, "2^32 (константа сети)", 4294967296, comment="Константа Bitcoin")
    put(ws, 22, "Награда блока, BTC", 3.125, comment="После халвинга 2024")
    put(ws, 23, "Эл-во в USD", f"={B('rub')}/{B('rate')}", comment="₽ за свет ÷ курс", fmt="0.00")
    put(ws, 24, "Комиссия партнёру, USD", f"={B('given')}*{B('price')}", comment="Отдано BTC × цена", fmt="0.00")
    put(ws, 25, "Все расходы, USD", f"={B('elec_usd')}+{B('comm_usd')}", comment="Эл-во $ + комиссия $", fmt="0.00")
    put(ws, 26, "Все расходы, ₽", f"={B('rub')}+{B('comm_usd')}*{B('rate')}", comment="Эл-во ₽ + комиссия в ₽", fmt="#,##0")
    put(ws, 27, "Намайнено NET, BTC", f"={B('btc')}-{B('given')}", comment="Gross − партнёру", fmt="0.00000")
    put(ws, 28, "HODL остаток, BTC", f"=MAX(0,{B('net_btc')}-{B('sold')})", comment="Net − продано", fmt="0.00000")
    put(ws, 29, "Теор. добыча, BTC",
        f"={B('hash')}*{B('days')}*86400*{B('reward')}/({B('diff')}*{B('pow32')})",
        comment="При 100% uptime. hash×сек×3.125/(diff×2³²)", fmt="0.00000")

    section(ws, 30, "💰 ФИНАНСЫ")
    put(ws, 31, "Доход от продаж, ₽", f"={B('sold')}*{B('price')}*{B('rate')}", comment="Продано × цена × курс", fmt="#,##0")
    put(ws, 32, "Стоимость HODL сейчас, ₽", f"={B('held')}*{B('cur_btc')}*{B('cur_rub')}", comment="Остаток × тек. цена BTC", fmt="#,##0")
    put(ws, 33, "Доход MTM (продажи+HODL), ₽", f"={B('rev_rub')}+{B('hodl_rub')}", comment="Полный доход по тек. цене", fmt="#,##0")
    put(ws, 34, "Реализованный кэш, ₽", f"={B('rev_rub')}-{B('cost_rub')}", comment="Только продажи − расходы", fmt="#,##0")
    put(ws, 35, "★ Фин. итог, ₽", f"={B('rev_rub')}+{B('hodl_rub')}-{B('cost_rub')}", comment="Продажи + HODL − расходы", fmt="#,##0")
    put(ws, 36, "Себестоимость 1 BTC, $", f"=IF({B('net_btc')}>0,{B('cost_usd')}/{B('net_btc')},0)", comment="All-in $/BTC", fmt="#,##0")
    put(ws, 37, "Себестоимость 1 BTC, ₽", f"={B('cost_btc_usd')}*{B('rate')}", comment="$/BTC × курс", fmt="#,##0")
    put(ws, 38, "Тариф, $/кВт·ч", f"=IF({B('kwh')}>0,{B('elec_usd')}/{B('kwh')},0)", comment="Стоимость 1 кВт·ч", fmt="0.0000")

    section(ws, 39, "🏭 KPI ЭФФЕКТИВНОСТИ")
    put(ws, 40, "J/TH", f"=IF({B('hash')}>0,{B('kwh')}*1000/({B('hash')}*{B('hours')}),0)",
        comment="<20 отлично (S21). >30 устарело", fmt="0.00")
    put(ws, 41, "Realization rate, %", f"=IF({B('theor')}>0,{B('btc')}/{B('theor')}*100,0)",
        comment="Факт/теор. 95-102% = норма", fmt="0.0")
    put(ws, 42, "Mining margin, %", f"=IF({B('mtm')}>0,({B('mtm')}-{B('cost_rub')})/{B('mtm')}*100,-100)",
        comment="(Доход MTM − расходы) / доход", fmt="0.0")
    put(ws, 43, "Break-even BTC, $", f"={B('cost_btc_usd')}", comment="= себестоимость $/BTC", fmt="#,##0")
    put(ws, 44, "Наш hashprice, $/TH/день", f"=IF({B('hash')}>0,{B('btc')}*{B('price')}/({B('hash')}*{B('days')}),0)",
        comment="Факт $/TH/day", fmt="0.0000")
    put(ws, 45, "Сетевой hashprice, $/TH/день",
        f"={B('reward')}*144*{B('price')}/({B('diff')}*{B('pow32')}/600)",
        comment="Бенчмарк всей сети", fmt="0.0000")
    put(ws, 46, "Hashprice: наш vs сеть, %",
        f"=IF({B('hp_net')}>0,({B('hp_our')}/{B('hp_net')}-1)*100,0)",
        comment="+% лучше сети, −% хуже", fmt="0.0")


def build_reference(ws):
    ws.title = "Справочник формул"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 48
    rows = [
        ("Показатель", "Формула", "От чего зависит"),
        ("NET BTC", "btcMined − btcGiven", "Gross и комиссия Демы"),
        ("Теор. BTC", "hash × дни × 86400 × 3.125 / (diff × 2³²)", "Hashrate, сложность, дни. НЕ от цены BTC"),
        ("Realization %", "btcMined / Теор. × 100", "<90% простои. >105% занизил hashrate"),
        ("J/TH", "кВт·ч × 1000 / (hash × часы)", "Потребление и hashrate"),
        ("All-in $/BTC", "(эл-во$ + комиссия$) / net BTC", "Все расходы на 1 BTC"),
        ("Фин. итог ₽", "продажи₽ + HODL₽ − расходы₽", "HODL = остаток × ТЕКУЩАЯ цена BTC"),
        ("Реализ. кэш", "продажи₽ − расходы₽", "Без HODL — только наличка"),
        ("Mining margin", "(MTM − расходы) / MTM", "Доля прибыли"),
        ("Hashprice наш", "btcMined × цена / (hash × дни)", "$/TH/день факт"),
        ("Hashprice сеть", "3.125 × 144 × цена / (diff × 2³² / 600)", "Среднее по сети"),
    ]
    for i, row in enumerate(rows, 1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(i, j).border = BORDER
        if i == 1:
            for j in range(1, 4):
                ws.cell(i, j).font = Font(bold=True)

    r = len(rows) + 2
    ws.cell(r, 1, "ЦЕПОЧКА: Ввод → Теор.BTC → Realization | кВт·ч+hash → J/TH | Расходы → $/BTC | Продажи+HODL → Фин.итог")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)


def build_examples(ws):
    ws.title = "Примеры (твои данные)"
    hdr = ["Локация", "Месяц", "кВт·ч", "Эл-во ₽", "TH/s", "Diff", "BTC", "Отдано", "Продано", "Цена$", "Курс", "Себест$/BTC", "Realiz%", "J/TH"]
    for i, h in enumerate(hdr, 1):
        ws.cell(1, i, h).font = Font(bold=True)
    data = [
        ["Бирск", "2026-04", 263325, 1316625, 20100, 135.59, 0.2741, 0, 0.15, 69000, 78, 61583, 98.0, 18.20],
        ["Дема", "2026-04", 89680, 325650, 4500, 135.59, 0.06174, 0, 0.01, 69000, 79, 66766, 98.6, 27.68],
        ["Бирск", "2026-05", 270000, 1351800, 19000, 136.61, 0.2777, 0, 0.2, 75000, 75, 64905, 102.5, 19.10],
        ["Дема", "2026-05", 64500, 310650, 4500, 136.61, 0.059716, 0, 0.01, 72000, 76, 68449, 93.0, 19.27],
    ]
    for ri, row in enumerate(data, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(ri, ci, v)


def main():
    wb = Workbook()
    build_calculator(wb.active)
    build_reference(wb.create_sheet())
    build_examples(wb.create_sheet())
    wb.save(OUT)
    import shutil
    shutil.copy(OUT, "/Users/kiber/Downloads/mining_calculator_formulas.xlsx")
    print("OK:", OUT)


if __name__ == "__main__":
    main()
